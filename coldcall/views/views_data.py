from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views import View
from django.views.generic import TemplateView

from django.utils.timezone import is_aware
from datetime import datetime, timezone
from dateutil.parser import parse

from .view_helper import get_template_dir, STUDENT_ATTRIBUTES, RATING_ATTRIBUTES
from ..models import Class, Student, StudentRating

import csv
import io
from zipfile import ZipFile

from openpyxl import Workbook

#Adds a list of students to a given class using a user provided .csv file
class AddStudentImportView(LoginRequiredMixin, TemplateView):

    def get(self, request, class_id=None): 
        self.template_name = get_template_dir("add_student_import", request.is_mobile)
        if class_id: 
            selected_class = Class.objects.get(id=class_id)
        else: 
            selected_class = None
        classes = Class.objects.filter(professor_key = request.user, is_archived=False)
        return render(request, self.template_name, {'classes': classes, 'selected_class': selected_class})
    
    def post(self, request, class_id=None):
        self.template_name = get_template_dir("add_student_import", request.is_mobile) 
        class_id = request.POST.get('class_id')
        student_file = request.FILES.get('students')
        rating_file = request.FILES.get("ratings")
        
        if not student_file:
            messages.error(request, "No student file uploaded. Please select a CSV file.")
            return render(request, self.template_name, {'classes': Class.objects.filter(professor_key=request.user), 'selected_class': None})
        
        # Check if the file type is csv
        if not student_file.name.endswith('.csv'):
            messages.error(request, "Invalid file type. Please select a CSV file.")
            return render(request, self.template_name, {'classes': Class.objects.filter(professor_key=request.user), 'selected_class': selected_class})
                
        try:
            selected_class = Class.objects.get(id=class_id)
        except Class.DoesNotExist:
            messages.error(request, "Invalid class ID. Please select a valid class for student import.")
            return render(request, self.template_name, {'classes': Class.objects.filter(professor_key=request.user), 'selected_class': None})

        try:
            decoded_file = student_file.read().decode('utf-8').splitlines()
            
            # Skip empty files
            if not decoded_file:
                messages.error(request, "The uploaded file is empty.")
                return render(request, self.template_name, {'classes': Class.objects.filter(professor_key=request.user), 'selected_class': selected_class})
            
            # Check if first row is header row
            is_header_row = False
            if decoded_file[0].lower().startswith('usc_id') or 'first_name' in decoded_file[0].lower():
                is_header_row = True
            
            # Simple CSV parsing approach
            rows_to_process = []
            expected_headers = ['usc_id', 'email', 'first_name', 'last_name', 'seating', 
                              'total_calls', 'absent_calls', 'total_score', 'class_id']
            
            for i, row in enumerate(decoded_file):
                # Skip header row
                if i == 0 and is_header_row:
                    continue
                
                if not row.strip():
                    continue
                
                # Add row to processing list
                rows_to_process.append(row)
            
            # Reading CSV file & gathering student information 
            successful_imports = 0
            
            for row in rows_to_process:
                # Split the CSV row
                values = row.split(',')
                
                # Create a dict with field names
                row_data = {}
                for i, value in enumerate(values):
                    if i < len(expected_headers):
                        row_data[expected_headers[i]] = value.strip()
                    else:
                        break
                
                # Extract fields with defaults
                usc_id = row_data.get('usc_id', '')
                email = row_data.get('email', '')
                first_name = row_data.get('first_name', '')
                last_name = row_data.get('last_name', '')
                seating = row_data.get('seating', 'NA')
                total_calls = row_data.get('total_calls', '0')
                absent_calls = row_data.get('absent_calls', '0')
                total_score = row_data.get('total_score', '0')
                
                # Skip rows that don't have the required fields
                if not usc_id or not email or not first_name or not last_name:
                    continue
                
                # Also skip if this looks like a header row
                if usc_id.lower() == 'usc_id' or 'first_name' in usc_id.lower():
                    continue
                
                try:
                    # Create or update the student
                    Student.objects.update_or_create(
                        usc_id=usc_id,
                        class_key=selected_class,
                        defaults={
                            'email': email, 
                            'first_name': first_name, 
                            'last_name': last_name,
                            'class_key': selected_class,
                            'seating': seating if seating in ['FR', "Front Right", 'FM', "Front Middle", 'FL', "Front Left", 'BR', "Back Right", 'BM', "Back Middle", 'BL', "Back Left", 'NA', "None"] else 'NA',
                            'total_calls': int(total_calls) if total_calls.isdigit() else 0,
                            'absent_calls': int(absent_calls) if absent_calls.isdigit() else 0,
                            'total_score': int(total_score) if total_score.isdigit() else 0
                        }
                    )
                    successful_imports += 1
                except Exception as e:
                    # Log the error but continue with other students
                    messages.warning(request, f"Error importing student {usc_id}: {str(e)}")
                    continue

            if successful_imports > 0:
                messages.success(request, f"Successfully imported {successful_imports} students.")
            else:
                messages.warning(request, "No students were imported. Please check your file format.")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return render(request, self.template_name, {'classes': Class.objects.filter(professor_key=request.user), 'selected_class': selected_class})

        # Handle rating import if file is present    
        if rating_file:
            if not rating_file.name.endswith('.csv'):
                messages.error(request, "Invalid rating file type. Please select a CSV file.")
                return redirect('/')
                
            try:
                decoded_file = rating_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)

                rows = []
                student_ids = set()        

                for row in reader:
                    usc_id = row.get('usc_id', '').strip()
                    date = row.get('date', '').strip()
                    #ensure valid data is present for key, skip otherwise
                    if not usc_id or not date:
                        continue
                    rows.append(row)
                    student_ids.add(usc_id)
                #bulk grab students using SQL IN search
                students = {
                    student.usc_id: student for student in Student.objects.filter(usc_id__in=student_ids)
                }
                #use new set to store student references rather than ids
                updated_students = set()
                #combine each query into one all-or-nothing query
                with transaction.atomic():
                    for row in rows:
                        usc_id = row.get('usc_id', '').strip()
                        date = row.get('date', '').strip()
                        attendance = row.get('attendance', "TRUE").strip().upper()
                        prepared = row.get('prepared', "FALSE").strip().upper()
                        score = row.get('score', '0').strip()

                        try:
                            date = parse(date)
                            if not date.tzinfo:
                                date = date.replace(tzinfo=timezone.utc)
                        except Exception:
                            continue

                        #grab student from existing set
                        student = students.get(usc_id)
                        if not student:
                            continue

                        StudentRating.objects.update_or_create(
                            student_key = student,
                            date = date,
                            defaults={
                                'attendance': attendance == "TRUE",
                                'prepared': prepared == "TRUE",
                                'score': int(score) if score.isdigit() else 0
                            }
                        )

                        updated_students.add(student)

                with transaction.atomic():
                    for student in updated_students:
                        student.recalculate_all()
            except Exception as e:
                messages.error(request, f"Error importing ratings: {str(e)}")
                
        return redirect('/')

class ExportClassFileView(View):
    def get(self, request, class_id=None):
        self.template_name = get_template_dir("export_class_file", request.is_mobile)
        classes = Class.objects.filter(professor_key=request.user)
        return render(request, self.template_name, {'classes': classes, 'class_id': class_id})
    
    def post(self, request, class_id=None):
        self.template_name = get_template_dir("export_class_file", request.is_mobile)
        class_ids = request.POST.getlist('class_id')
        export_type = request.POST.get('export_type')
        file_format = request.POST.get('file_format', 'csv')

        if not class_ids or not export_type:
            messages.error(request, "No class ID provided.")
            return render(request, self.template_name, {'classes': Class.objects.filter(professor_key=request.user), 'class_id': class_id})
        
        try:
            content_types = {
                'csv': 'text/csv',
                'txt': 'text/plain',
                'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            extensions = {'csv': '.csv', 'txt': '.txt', 'excel': '.xlsx'}
            content_type = content_types.get(file_format, 'text/csv')
            extension = extensions.get(file_format, '.csv')

            def format_iso_date(date):
                if isinstance(date, datetime):
                    if is_aware(date):
                        return date.isoformat()
                    return date.replace(tzinfo=None).isoformat
                return date

        
            # Generates the rows for the given class
            # Simple only exports a list of student attributes, otherwise export student ratings
            def generate_file(class_obj):
                rows = [STUDENT_ATTRIBUTES] if export_type == "simple" else [RATING_ATTRIBUTES]
                for student in class_obj.student_set.all():
                    if export_type == "simple":
                        rows.append(student.get_data_list())
                    else:
                        for rating in student.studentrating_set.all():
                            rows.append([student.usc_id, 
                                         format_iso_date(rating.date), 
                                         rating.attendance, rating.prepared, rating.score, student.class_key.pk])

                return rows
        
            
            response = HttpResponse(content_type=content_type)
            #export a single file, not to zip
            if len(class_ids) == 1:
                class_obj = Class.objects.get(id=class_ids[0], professor_key=request.user)
                filename = f"{class_obj.class_name}_{datetime.now().strftime('%Y%m%d')}{"" if export_type == "simple" else "_ratings"}{extension}"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'

                rows = generate_file(class_obj)
                # CSV/txt and Excel require different writers
                if file_format in ['csv', 'txt']:
                    writer = csv.writer(response)
                    writer.writerows(rows)
                else:
                    wb = Workbook() 
                    ws = wb.active
                    for row in rows:
                        ws.append(row)
                    wb.save(response)

                return response
            #iterate through each class id, create file, and add to .zip
            zip_buffer = io.BytesIO()
            with ZipFile(zip_buffer, 'w') as zip_file:
                for class_id in class_ids:
                    class_obj = Class.objects.get(id=class_id, professor_key=request.user)

                    rows = generate_file(class_obj)

                    if file_format in ['csv', 'txt']:
                        output = io.StringIO()
                        writer = csv.writer(output)
                        writer.writerows(rows)
                        content = output.getvalue()
                    else: #Excel
                        output = io.BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        for row in rows: 
                            ws.append(row)
                        wb.save(output)
                        content = output.getvalue()

                    timestamp = datetime.now().strftime('%Y%m%d')
                    filename = f"{class_obj.class_name}_{timestamp}{"" if export_type == "simple" else "_ratings"}{extension}"
                    zip_file.writestr(filename, content)

            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="class_exports_{datetime.now().strftime('%Y%m%d')}.zip"'
            
            return response
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return render(request, self.template_name, {'classes': Class.objects.filter(professor_key=request.user), 'class_id': class_id})

# TODO: Test above to ensure it functions the same, then remove commented section
# Gives the user a .csv file containing information about each student in a class
# class ExportClassFileView2(View):
#     def get(self, request):
#         self.template_name = get_template_dir("export_class_file", request.is_mobile)
#         classes = Class.objects.filter(professor_key=request.user)
#         return render(request, self.template_name, {'classes': classes})
    
#     def post(self, request):
#         class_ids = request.POST.getlist('class_id')
#         export_type = request.POST.get('export_type')
#         file_format = request.POST.get('file_format', 'csv')
        
#         if not class_ids or not export_type:
#             return HttpResponseBadRequest("No class ID provided.")

#         try:
#             # Set content type and extension based on format
#             content_types = {
#                 'csv': 'text/csv',
#                 'txt': 'text/plain',
#                 'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             }
#             extensions = {'csv': '.csv', 'txt': '.txt', 'excel': '.xlsx'}
#             content_type = content_types.get(file_format, 'text/csv')
#             extension = extensions.get(file_format, '.csv')

#             # If only one class is selected, return a single file
#             if len(class_ids) == 1:
#                 class_obj = Class.objects.get(id=class_ids[0], professor_key=request.user)
#                 timestamp = datetime.now().strftime('%Y%m%d')
#                 filename = f"{class_obj.class_name}_{timestamp}{extension}"

#                 if file_format in ['csv', 'txt']:
#                     response = HttpResponse(content_type=content_type)
#                     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#                     writer = csv.writer(response)
                    
#                     if export_type == "simple":
#                         writer.writerow(['usc_id', 'email', 'first_name', 'last_name', 'seating', 'total_calls', 'absent_calls', 'total_score', 'class_id'])
#                         for student in class_obj.student_set.all():
#                             writer.writerow([student.usc_id, student.email, student.first_name, student.last_name,
#                                           student.seating, student.total_calls, student.absent_calls, student.total_score, student.class_key.pk])
#                     else:  # export_type == "all"
#                         writer.writerow(['usc_id', 'date', 'attendance', 'prepared', 'score', 'class_id'])
#                         for student in class_obj.student_set.all():
#                             ratings = StudentRating.objects.filter(student_key=student)
#                             for rating in ratings:
#                                 writer.writerow([student.usc_id, rating.date, rating.attendance,
#                                                rating.prepared, rating.score, student.class_key.pk])
#                 else:  # Excel format
#                     wb = Workbook()
#                     ws = wb.active
                    
#                     if export_type == "simple":
#                         ws.append(['usc_id', 'email', 'first_name', 'last_name', 'seating', 'total_calls', 'absent_calls', 'total_score', 'class_id'])
#                         for student in class_obj.student_set.all():
#                             ws.append([student.usc_id, student.email, student.first_name, student.last_name,
#                                      student.seating, student.total_calls, student.absent_calls, student.total_score, student.class_key.pk])
#                     else:  # export_type == "all"
#                         ws.append(['usc_id', 'date', 'attendance', 'prepared', 'score', 'class_id'])
#                         for student in class_obj.student_set.all():
#                             ratings = StudentRating.objects.filter(student_key=student)
#                             for rating in ratings:
#                                 ws.append([student.usc_id, rating.date, rating.attendance,
#                                          rating.prepared, rating.score, student.class_key.pk])
                    
#                     response = HttpResponse(content_type=content_type)
#                     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#                     wb.save(response)

#                 return response

#             # Multiple classes selected - create ZIP file
#             else:
#                 zip_buffer = io.BytesIO()
#                 with ZipFile(zip_buffer, 'w') as zip_file:
#                     for class_id in class_ids:
#                         class_obj = Class.objects.get(id=class_id, professor_key=request.user)
                        
#                         if file_format in ['csv', 'txt']:
#                             output = io.StringIO()
#                             writer = csv.writer(output)
                            
#                             if export_type == "simple":
#                                 writer.writerow(['usc_id', 'email', 'first_name', 'last_name', 'seating', 'total_calls', 'absent_calls', 'total_score', 'class_id'])
#                                 for student in class_obj.student_set.all():
#                                     writer.writerow([student.usc_id, student.email, student.first_name, student.last_name,
#                                                   student.seating, student.total_calls, student.absent_calls, student.total_score, student.class_key.pk])
#                             else:  # export_type == "all"
#                                 writer.writerow(['usc_id', 'date', 'attendance', 'prepared', 'score', 'class_id'])
#                                 for student in class_obj.student_set.all():
#                                     ratings = StudentRating.objects.filter(student_key=student)
#                                     for rating in ratings:
#                                         writer.writerow([student.usc_id, rating.date, rating.attendance,
#                                                        rating.prepared, rating.score, student.class_key.pk])
                            
#                             content = output.getvalue()
                            
#                         else:  # Excel format
#                             output = io.BytesIO()
#                             wb = Workbook()
#                             ws = wb.active
                            
#                             if export_type == "simple":
#                                 ws.append(['usc_id', 'email', 'first_name', 'last_name', 'seating', 'total_calls', 'absent_calls', 'total_score', 'class_id'])
#                                 for student in class_obj.student_set.all():
#                                     ws.append([student.usc_id, student.email, student.first_name, student.last_name,
#                                              student.seating, student.total_calls, student.absent_calls, student.total_score, student.class_key.pk])
#                             else:  # export_type == "all"
#                                 ws.append(['usc_id', 'date', 'attendance', 'prepared', 'score', 'class_id'])
#                                 for student in class_obj.student_set.all():
#                                     ratings = StudentRating.objects.filter(student_key=student)
#                                     for rating in ratings:
#                                         ws.append([student.usc_id, rating.date, rating.attendance,
#                                                  rating.prepared, rating.score, student.class_key.pk])
                            
#                             wb.save(output)
#                             content = output.getvalue()

#                         timestamp = datetime.now().strftime('%Y%m%d')
#                         filename = f"{class_obj.class_name}_{timestamp}{extension}"
#                         zip_file.writestr(filename, content)

#                 response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
#                 response['Content-Disposition'] = f'attachment; filename="class_exports.zip"'
                
#                 return response
            
#         except Exception as e:
#             return HttpResponseBadRequest(f"Error: {str(e)}")